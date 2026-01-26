import pandas as pd
from openpyxl import load_workbook
import json
import warnings
warnings.filterwarnings('ignore')

file_path = "/mnt/user-data/uploads/프로젝트_희귀분석___.xlsm"

# 1. Raw_Retention 데이터 추출
print("Extracting Raw_Retention...")
wb = load_workbook(file_path, data_only=True)
ws = wb['Raw_Retention']

retention_data = {}
for row in range(4, 27):  # 게임 데이터 행
    game_name = ws.cell(row=row, column=2).value
    if game_name and game_name != '게임명' and not str(game_name).startswith('-'):
        retention_values = []
        for col in range(3, 95):  # D+1 ~ D+90
            val = ws.cell(row=row, column=col).value
            if val is not None and isinstance(val, (int, float)) and val < 1:  # 리텐션은 1 미만
                retention_values.append(round(float(val), 6))
            else:
                break
        if retention_values:
            retention_data[game_name] = retention_values

print(f"  - {len(retention_data)} games extracted")

# 2. Raw_NRU 데이터 추출
print("Extracting Raw_NRU...")
ws = wb['Raw_NRU']

nru_data = {}
for row in range(4, 28):
    game_name = ws.cell(row=row, column=2).value
    if game_name and game_name != '게임명' and not str(game_name).startswith('-'):
        nru_values = []
        for col in range(3, 370):  # D+1 ~ D+365
            val = ws.cell(row=row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                nru_values.append(int(val))
            else:
                nru_values.append(0)
        # 끝의 0 제거
        while nru_values and nru_values[-1] == 0:
            nru_values.pop()
        if nru_values:
            nru_data[game_name] = nru_values

print(f"  - {len(nru_data)} games extracted")

# 3. Raw_PR (Payment Rate) 데이터 추출
print("Extracting Raw_PR...")
ws = wb['Raw_PR']

pr_data = {}
for row in range(4, 28):
    game_name = ws.cell(row=row, column=2).value
    if game_name and game_name != '게임명' and not str(game_name).startswith('-'):
        pr_values = []
        for col in range(3, 370):
            val = ws.cell(row=row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                pr_values.append(round(float(val), 6))
            else:
                pr_values.append(0)
        while pr_values and pr_values[-1] == 0:
            pr_values.pop()
        if pr_values:
            pr_data[game_name] = pr_values

print(f"  - {len(pr_data)} games extracted")

# 4. Raw_ARPPU 데이터 추출
print("Extracting Raw_ARPPU...")
ws = wb['Raw_ARPPU']

arppu_data = {}
for row in range(4, 28):
    game_name = ws.cell(row=row, column=2).value
    if game_name and game_name != '게임명' and not str(game_name).startswith('-'):
        arppu_values = []
        for col in range(3, 370):
            val = ws.cell(row=row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                arppu_values.append(round(float(val), 2))
            else:
                arppu_values.append(0)
        while arppu_values and arppu_values[-1] == 0:
            arppu_values.pop()
        if arppu_values:
            arppu_data[game_name] = arppu_values

print(f"  - {len(arppu_data)} games extracted")

wb.close()

# JSON 스키마로 통합
raw_data = {
    "version": "1.0",
    "description": "Game KPI Raw Data for Projection",
    "games": {
        "retention": retention_data,
        "nru": nru_data,
        "payment_rate": pr_data,
        "arppu": arppu_data
    },
    "metadata": {
        "retention_games": list(retention_data.keys()),
        "nru_games": list(nru_data.keys()),
        "pr_games": list(pr_data.keys()),
        "arppu_games": list(arppu_data.keys())
    }
}

# JSON 파일로 저장
with open('/home/claude/game-kpi-projection/data/raw_game_data.json', 'w', encoding='utf-8') as f:
    json.dump(raw_data, f, ensure_ascii=False, indent=2)

print("\n✅ JSON export completed: data/raw_game_data.json")
print(f"\nSummary:")
print(f"  - Retention: {len(retention_data)} games")
print(f"  - NRU: {len(nru_data)} games")
print(f"  - Payment Rate: {len(pr_data)} games")
print(f"  - ARPPU: {len(arppu_data)} games")
