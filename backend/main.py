from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Dict, Optional, Any
import numpy as np
from scipy.optimize import curve_fit
import json
import os
import httpx

app = FastAPI(title="Game KPI Projection API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Data paths
DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
RAW_DATA_PATH = os.path.join(DATA_DIR, "raw_game_data.json")
CONFIG_PATH = os.path.join(DATA_DIR, "default_config.json")

# Claude API Configuration
# API Keys (환경변수에서만 읽어옴 - 코드에 키 포함 금지!)
CLAUDE_API_KEY = os.environ.get("CLAUDE_API_KEY", "")
CLAUDE_API_URL = "https://api.anthropic.com/v1/messages"

def load_raw_data():
    with open(RAW_DATA_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def load_config():
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

# Pydantic Models
class RetentionInput(BaseModel):
    selected_games: List[str] = []
    target_d1_retention: Dict[str, float] = {"best": 0.45, "normal": 0.40, "worst": 0.35}

class NRUInput(BaseModel):
    selected_games: List[str] = []
    d1_nru: Dict[str, int] = {"best": 0, "normal": 0, "worst": 0}
    paid_organic_ratio: float = 0.5
    nvr: float = 0.7
    adjustment: Dict[str, float] = {"best_vs_normal": 0.1, "worst_vs_normal": -0.1}
    # V8.5: UA/Brand 예산 분리
    ua_budget: Optional[int] = 0              # 퍼포먼스 마케팅 예산 (직접 유입)
    brand_budget: Optional[int] = 0           # 브랜딩 예산 (Organic Boost)
    target_cpa: Optional[int] = 2000          # CPI/CPA (ua_budget에만 적용)
    base_organic_ratio: Optional[float] = 0.2 # 기본 자연 유입 비율
    # V8.5+: Pre-Launch & CPA Saturation
    pre_marketing_ratio: Optional[float] = 0.0    # 사전 마케팅 비중 (0~1, 예: 0.3 = 30%)
    wishlist_conversion_rate: Optional[float] = 0.15  # 위시리스트/사전예약 → 실제 유입 전환율 (PC: 10~20%)
    cpa_saturation_enabled: Optional[bool] = True     # CPA 상승 계수 활성화
    brand_time_lag_enabled: Optional[bool] = True     # 브랜딩 지연 효과 활성화

class RevenueInput(BaseModel):
    selected_games_pr: List[str] = []
    selected_games_arppu: List[str] = []
    pr_adjustment: Dict[str, float] = {"best_vs_normal": 0.05, "worst_vs_normal": -0.05}
    arppu_adjustment: Dict[str, float] = {"best_vs_normal": 0.05, "worst_vs_normal": -0.05}

class ProjectionInput(BaseModel):
    launch_date: str
    projection_days: int = 365
    retention: RetentionInput
    nru: NRUInput
    revenue: RevenueInput
    basic_settings: Optional[Dict[str, Any]] = None
    # 블렌딩 설정
    blending: Optional[Dict[str, Any]] = None  # { weight: 0.7, genre: "MMORPG", platforms: ["PC"] }
    # V7 추가: 품질 점수, BM 타입, 지역
    quality_score: Optional[str] = "B"  # S/A/B/C/D
    bm_type: Optional[str] = "Midcore"  # Hardcore/Midcore/Casual/F2P_Cosmetic/Gacha
    regions: Optional[List[str]] = None  # ["korea", "japan", "global", ...]

# ============================================
# 글로벌 계절성 팩터 (지역별 월간 가중치)
# ============================================
SEASONALITY_BY_REGION = {
    "korea": {1: 1.15, 2: 1.20, 3: 1.00, 4: 0.95, 5: 1.00, 6: 0.95, 7: 1.05, 8: 1.10, 9: 1.00, 10: 1.05, 11: 1.10, 12: 1.15},
    "japan": {1: 1.10, 2: 1.05, 3: 1.05, 4: 1.10, 5: 1.15, 6: 0.95, 7: 1.00, 8: 1.05, 9: 1.00, 10: 1.00, 11: 1.05, 12: 1.20},
    "china": {1: 1.10, 2: 1.25, 3: 1.00, 4: 0.95, 5: 1.05, 6: 1.10, 7: 1.05, 8: 1.00, 9: 1.00, 10: 1.20, 11: 1.15, 12: 1.05},
    "global": {1: 0.95, 2: 0.90, 3: 0.95, 4: 1.00, 5: 1.00, 6: 1.00, 7: 1.05, 8: 1.00, 9: 1.00, 10: 1.05, 11: 1.15, 12: 1.25},
    "sea": {1: 1.05, 2: 1.10, 3: 1.00, 4: 1.00, 5: 1.00, 6: 1.05, 7: 1.05, 8: 1.00, 9: 1.00, 10: 1.00, 11: 1.05, 12: 1.15},
    "na": {1: 0.90, 2: 0.90, 3: 0.95, 4: 1.00, 5: 1.00, 6: 1.05, 7: 1.05, 8: 1.00, 9: 0.95, 10: 1.05, 11: 1.20, 12: 1.25},
    "sa": {1: 1.10, 2: 1.05, 3: 1.00, 4: 0.95, 5: 0.95, 6: 1.00, 7: 1.05, 8: 1.00, 9: 1.00, 10: 1.05, 11: 1.10, 12: 1.15},
    "eu": {1: 0.90, 2: 0.90, 3: 0.95, 4: 1.05, 5: 1.00, 6: 1.00, 7: 1.00, 8: 0.95, 9: 1.00, 10: 1.05, 11: 1.15, 12: 1.25},
}

def calculate_seasonality(regions: List[str], launch_date: str, days: int = 365) -> List[float]:
    """
    지역별 계절성 팩터 계산
    - 월간 기본 계절성
    - 주간 변동성 (주말 +15~20%)
    - 특별 이벤트 스파이크 (명절, 대형 업데이트 등)
    """
    from datetime import datetime, timedelta
    import random
    
    try:
        start_date = datetime.strptime(launch_date, "%Y-%m-%d")
    except:
        start_date = datetime(2026, 11, 12)  # 기본값
    
    # 시드 고정 (재현성)
    random.seed(42)
    
    # 특별 이벤트 날짜 (월-일 기준)
    SPECIAL_EVENTS = {
        "korea": [(1, 1), (2, 1), (2, 2), (5, 5), (9, 15), (9, 16), (9, 17), (12, 25), (12, 31)],  # 설날, 추석, 크리스마스 등
        "japan": [(1, 1), (5, 3), (5, 4), (5, 5), (8, 15), (12, 25), (12, 31)],  # 신정, 골든위크, 오본 등
        "global": [(1, 1), (11, 24), (11, 25), (12, 24), (12, 25), (12, 31)],  # 블랙프라이데이, 크리스마스 등
        "na": [(1, 1), (7, 4), (11, 24), (11, 25), (12, 24), (12, 25), (12, 31)],
        "eu": [(1, 1), (12, 24), (12, 25), (12, 31)],
        "china": [(1, 1), (2, 1), (2, 2), (10, 1), (10, 2), (10, 3)],  # 춘절, 국경절
        "sea": [(1, 1), (4, 13), (4, 14), (11, 1), (12, 25), (12, 31)],  # 송끄란 등
        "sa": [(1, 1), (2, 13), (2, 14), (12, 25), (12, 31)],  # 카니발 등
    }
    
    factors = []
    for day in range(days):
        current_date = start_date + timedelta(days=day)
        month = current_date.month
        weekday = current_date.weekday()  # 0=월, 6=일
        month_day = (current_date.month, current_date.day)
        
        # 1. 월간 기본 계절성
        region_factors = []
        for region in regions:
            region_key = region.lower()
            if region_key in SEASONALITY_BY_REGION:
                region_factors.append(SEASONALITY_BY_REGION[region_key].get(month, 1.0))
        
        base_factor = np.mean(region_factors) if region_factors else 1.0
        
        # 2. 주간 변동성 (금~일 +15~20%, 월~화 -5~10%)
        if weekday == 4:  # 금요일
            weekly_factor = 1.12 + random.uniform(0, 0.05)
        elif weekday == 5:  # 토요일
            weekly_factor = 1.18 + random.uniform(0, 0.07)
        elif weekday == 6:  # 일요일
            weekly_factor = 1.15 + random.uniform(0, 0.05)
        elif weekday in [0, 1]:  # 월/화
            weekly_factor = 0.92 + random.uniform(0, 0.05)
        else:  # 수/목
            weekly_factor = 1.0 + random.uniform(-0.02, 0.02)
        
        # 3. 특별 이벤트 스파이크 (+30~60%)
        event_factor = 1.0
        for region in regions:
            region_key = region.lower()
            if region_key in SPECIAL_EVENTS:
                if month_day in SPECIAL_EVENTS[region_key]:
                    event_factor = max(event_factor, 1.35 + random.uniform(0, 0.25))
        
        # 4. 대형 업데이트 시뮬레이션 (30일마다 +20~35%)
        if day > 30 and (day % 30 < 3 or day % 30 > 27):
            event_factor = max(event_factor, 1.20 + random.uniform(0, 0.15))
        
        # 5. 약간의 랜덤 노이즈 (±3%)
        noise = 1.0 + random.uniform(-0.03, 0.03)
        
        final_factor = base_factor * weekly_factor * event_factor * noise
        factors.append(final_factor)
    
    return factors

# ============================================
# Time-Decay 블렌딩 (시간에 따라 가중치 변경)
# ============================================
def calculate_time_decay_weight(day: int, days: int = 365) -> float:
    """
    시간 가중치 계산 (Time-Decay)
    
    D1: 내부 90% : 벤치마크 10%
    D180: 내부 50% : 벤치마크 50%
    D365: 내부 10% : 벤치마크 90%
    
    선형 보간으로 매일 가중치 변경
    """
    # D1 = 0.9, D365 = 0.1 (선형 감소)
    weight_internal = 0.9 - (0.8 * (day - 1) / (days - 1)) if days > 1 else 0.9
    return max(min(weight_internal, 0.9), 0.1)

def calculate_time_decay_blended_retention(
    internal_curve: List[float],
    benchmark_curve: List[float],
    days: int = 365,
    quality_score: float = 1.0
) -> List[float]:
    """
    Time-Decay 블렌딩 리텐션 커브 생성
    
    Args:
        internal_curve: 내부 표본 리텐션 커브
        benchmark_curve: 벤치마크 리텐션 커브
        days: 프로젝션 기간
        quality_score: 품질 점수 (S=1.2, A=1.1, B=1.0, C=0.9, D=0.8)
    """
    blended = []
    for day in range(days):
        weight_internal = calculate_time_decay_weight(day + 1, days)
        weight_benchmark = 1 - weight_internal
        
        internal_val = internal_curve[day] if day < len(internal_curve) else internal_curve[-1]
        benchmark_val = benchmark_curve[day] if day < len(benchmark_curve) else benchmark_curve[-1]
        
        # 벤치마크에 품질 점수 적용
        adjusted_benchmark = benchmark_val * quality_score
        
        blended_val = (internal_val * weight_internal) + (adjusted_benchmark * weight_benchmark)
        blended.append(max(min(blended_val, 1.0), 0.001))
    
    return blended

# ============================================
# Quality Score 정의
# ============================================
QUALITY_SCORES = {
    "S": 1.2,   # 최상급 (FGT/CBT 결과 매우 우수)
    "A": 1.1,   # 우수
    "B": 1.0,   # 보통 (기본값)
    "C": 0.9,   # 미흡
    "D": 0.8,   # 부진
}

# ============================================
# BM 타입별 세분화 벤치마크
# ============================================
BM_TYPE_MODIFIERS = {
    "Hardcore": {"pr_mod": 0.5, "arppu_mod": 2.0},    # 낮은 PR, 고액 ARPPU
    "Midcore": {"pr_mod": 1.0, "arppu_mod": 1.0},     # 기본
    "Casual": {"pr_mod": 2.0, "arppu_mod": 0.4},      # 높은 PR, 소액 ARPPU
    "F2P_Cosmetic": {"pr_mod": 0.8, "arppu_mod": 0.6}, # 무료+꾸미기 중심
    "Gacha": {"pr_mod": 1.5, "arppu_mod": 1.8},       # 가챠 중심
}
BENCHMARK_DATA = {
    "PC": {
        "MMORPG": {"d1": 0.32, "d7": 0.20, "d30": 0.11, "d90": 0.06, "pr": 0.06, "arppu": 78000},
        "Action RPG": {"d1": 0.30, "d7": 0.18, "d30": 0.09, "d90": 0.04, "pr": 0.05, "arppu": 65000},
        "Battle Royale": {"d1": 0.35, "d7": 0.22, "d30": 0.12, "d90": 0.07, "pr": 0.03, "arppu": 45000},
        "Extraction Shooter": {"d1": 0.28, "d7": 0.16, "d30": 0.08, "d90": 0.04, "pr": 0.04, "arppu": 55000},
        "FPS/TPS": {"d1": 0.33, "d7": 0.20, "d30": 0.10, "d90": 0.05, "pr": 0.04, "arppu": 50000},
        "Strategy": {"d1": 0.25, "d7": 0.15, "d30": 0.08, "d90": 0.04, "pr": 0.07, "arppu": 85000},
        "Casual": {"d1": 0.40, "d7": 0.20, "d30": 0.08, "d90": 0.03, "pr": 0.02, "arppu": 25000},
        "Sports": {"d1": 0.30, "d7": 0.18, "d30": 0.09, "d90": 0.04, "pr": 0.05, "arppu": 60000},
    },
    "Mobile": {
        "MMORPG": {"d1": 0.42, "d7": 0.18, "d30": 0.07, "d90": 0.03, "pr": 0.05, "arppu": 52000},
        "Action RPG": {"d1": 0.38, "d7": 0.15, "d30": 0.06, "d90": 0.02, "pr": 0.04, "arppu": 45000},
        "Battle Royale": {"d1": 0.45, "d7": 0.20, "d30": 0.08, "d90": 0.04, "pr": 0.02, "arppu": 35000},
        "Extraction Shooter": {"d1": 0.35, "d7": 0.14, "d30": 0.05, "d90": 0.02, "pr": 0.03, "arppu": 40000},
        "FPS/TPS": {"d1": 0.40, "d7": 0.17, "d30": 0.07, "d90": 0.03, "pr": 0.03, "arppu": 38000},
        "Strategy": {"d1": 0.35, "d7": 0.16, "d30": 0.07, "d90": 0.03, "pr": 0.06, "arppu": 68000},
        "Casual": {"d1": 0.50, "d7": 0.22, "d30": 0.09, "d90": 0.04, "pr": 0.02, "arppu": 18000},
        "Sports": {"d1": 0.38, "d7": 0.16, "d30": 0.06, "d90": 0.02, "pr": 0.04, "arppu": 42000},
    },
    "Console": {
        "MMORPG": {"d1": 0.35, "d7": 0.22, "d30": 0.12, "d90": 0.06, "pr": 0.05, "arppu": 70000},
        "Action RPG": {"d1": 0.33, "d7": 0.20, "d30": 0.10, "d90": 0.05, "pr": 0.04, "arppu": 60000},
        "Battle Royale": {"d1": 0.38, "d7": 0.24, "d30": 0.13, "d90": 0.07, "pr": 0.02, "arppu": 40000},
        "Extraction Shooter": {"d1": 0.30, "d7": 0.18, "d30": 0.09, "d90": 0.04, "pr": 0.03, "arppu": 50000},
        "FPS/TPS": {"d1": 0.36, "d7": 0.22, "d30": 0.11, "d90": 0.06, "pr": 0.03, "arppu": 48000},
        "Strategy": {"d1": 0.28, "d7": 0.17, "d30": 0.09, "d90": 0.04, "pr": 0.06, "arppu": 75000},
        "Casual": {"d1": 0.42, "d7": 0.20, "d30": 0.08, "d90": 0.03, "pr": 0.02, "arppu": 22000},
        "Sports": {"d1": 0.35, "d7": 0.20, "d30": 0.10, "d90": 0.05, "pr": 0.05, "arppu": 55000},
    }
}

def get_benchmark_data(genre: str, platforms: List[str]) -> Dict[str, float]:
    """장르/플랫폼에 맞는 벤치마크 데이터 반환 (다중 플랫폼은 평균)"""
    if not platforms:
        platforms = ["PC"]
    
    values = []
    for platform in platforms:
        if platform in BENCHMARK_DATA and genre in BENCHMARK_DATA[platform]:
            values.append(BENCHMARK_DATA[platform][genre])
    
    if not values:
        # 기본값 (PC/MMORPG)
        return {"d1": 0.32, "d7": 0.20, "d30": 0.11, "d90": 0.06, "pr": 0.06, "arppu": 78000}
    
    # 다중 플랫폼이면 평균
    return {
        "d1": np.mean([v["d1"] for v in values]),
        "d7": np.mean([v["d7"] for v in values]),
        "d30": np.mean([v["d30"] for v in values]),
        "d90": np.mean([v["d90"] for v in values]),
        "pr": np.mean([v["pr"] for v in values]),
        "arppu": np.mean([v["arppu"] for v in values]),
    }

def generate_benchmark_retention_curve(benchmark: Dict[str, float], days: int = 365) -> List[float]:
    """벤치마크 데이터로 Power Law 리텐션 커브 생성"""
    # D1, D7, D30, D90 데이터로 회귀분석
    x_data = np.array([1, 7, 30, 90])
    y_data = np.array([benchmark["d1"], benchmark["d7"], benchmark["d30"], benchmark["d90"]])
    
    try:
        popt, _ = curve_fit(retention_curve, x_data, y_data, p0=[0.5, -0.3], maxfev=5000)
        a, b = popt
    except:
        a, b = benchmark["d1"], -0.5  # 기본값
    
    curve = []
    for day in range(1, days + 1):
        ret = a * np.power(day, b)
        ret = max(min(ret, 1.0), 0.001)
        curve.append(ret)
    
    return curve

def calculate_blended_retention(
    internal_curve: List[float],
    benchmark_curve: List[float],
    weight_internal: float
) -> List[float]:
    """내부 표본과 벤치마크를 블렌딩한 리텐션 커브 생성"""
    weight_benchmark = 1 - weight_internal
    
    blended = []
    for i in range(len(internal_curve)):
        internal_val = internal_curve[i] if i < len(internal_curve) else internal_curve[-1]
        benchmark_val = benchmark_curve[i] if i < len(benchmark_curve) else benchmark_curve[-1]
        blended_val = (internal_val * weight_internal) + (benchmark_val * weight_benchmark)
        blended.append(max(min(blended_val, 1.0), 0.001))
    
    return blended

def calculate_blended_pr(
    internal_pr: List[float],
    benchmark_pr: float,
    weight_internal: float,
    days: int = 365
) -> List[float]:
    """PR 블렌딩"""
    weight_benchmark = 1 - weight_internal
    
    blended = []
    for i in range(days):
        internal_val = internal_pr[i] if i < len(internal_pr) else internal_pr[-1]
        blended_val = (internal_val * weight_internal) + (benchmark_pr * weight_benchmark)
        blended.append(max(min(blended_val, 1.0), 0.001))
    
    return blended

def calculate_blended_arppu(
    internal_arppu: List[float],
    benchmark_arppu: float,
    weight_internal: float,
    days: int = 365
) -> List[float]:
    """ARPPU 블렌딩"""
    weight_benchmark = 1 - weight_internal
    
    blended = []
    for i in range(days):
        internal_val = internal_arppu[i] if i < len(internal_arppu) else internal_arppu[-1]
        blended_val = (internal_val * weight_internal) + (benchmark_arppu * weight_benchmark)
        blended.append(max(blended_val, 1000))
    
    return blended

class AIInsightRequest(BaseModel):
    projection_summary: Dict[str, Any]
    analysis_type: str = "general"  # general, retention, nru, revenue, risk

# Retention Curve: a * (day)^b
def retention_curve(x, a, b):
    return a * np.power(x, b)

def fit_retention_curve(retention_data: List[float]):
    days = np.arange(1, len(retention_data) + 1)
    retention = np.array(retention_data)
    
    valid_mask = (retention > 0) & (retention <= 1)
    if np.sum(valid_mask) < 3:
        return None, None
    
    try:
        popt, _ = curve_fit(
            retention_curve, 
            days[valid_mask], 
            retention[valid_mask],
            p0=[retention_data[0], -0.5],
            bounds=([0, -2], [2, 0]),
            maxfev=5000
        )
        return popt[0], popt[1]
    except:
        return retention_data[0], -0.5

def calculate_retention_coefficients(selected_games: List[str], raw_data: dict):
    retention_games = raw_data['games']['retention']
    
    a_values = []
    b_values = []
    
    for game in selected_games:
        if game in retention_games:
            a, b = fit_retention_curve(retention_games[game])
            if a is not None:
                a_values.append(a)
                b_values.append(b)
    
    if not a_values:
        return 1.0, -0.5
    
    return np.mean(a_values), np.mean(b_values)

def generate_retention_curve(a: float, b: float, target_d1: float, days: int = 365):
    base_d1 = retention_curve(1, a, b)
    if base_d1 > 0:
        scale_factor = target_d1 / base_d1
    else:
        scale_factor = 1.0
    
    curve = []
    for day in range(1, days + 1):
        ret = retention_curve(day, a, b) * scale_factor
        curve.append(min(max(ret, 0.001), 1))
    
    return curve

def calculate_nru_pattern(selected_games: List[str], raw_data: dict):
    nru_games = raw_data['games']['nru']
    
    valid_games = [g for g in selected_games if g in nru_games]
    if not valid_games:
        return [0.98 ** i for i in range(365)]
    
    min_len = min(len(nru_games[g]) for g in valid_games)
    min_len = min(min_len, 365)
    
    daily_ratios = []
    for day in range(1, min_len):
        day_ratios = []
        for game in valid_games:
            data = nru_games[game]
            if day < len(data) and data[day-1] > 0:
                ratio = data[day] / data[day-1]
                if 0 < ratio < 2:
                    day_ratios.append(ratio)
        if day_ratios:
            daily_ratios.append(np.mean(day_ratios))
        else:
            daily_ratios.append(0.98)
    
    while len(daily_ratios) < 364:
        daily_ratios.append(daily_ratios[-1] if daily_ratios else 0.98)
    
    return daily_ratios

def generate_nru_series(total_nru: int, daily_ratios: List[float], days: int = 365, 
                         launch_period: int = 30, sustaining_ratio: float = 0.1):
    """
    NRU 시리즈 생성 - 런칭 마케팅은 D1~D30에 집중
    
    🔥 V8.3 수정: Area Normalization 적용
    - total_nru: 런칭 기간 동안의 "총 모객 수" (예산/CPI로 계산된 값)
    - 이 총량을 30일 패턴의 면적(Area)으로 나누어 D1 높이(Scale)를 산출
    - 결과: 예산 범위 내에서 유저가 분산 유입됨
    
    Args:
        total_nru: 런칭 기간 총 모객 수 (🔥 기존 d1_nru → total_nru로 해석 변경)
        daily_ratios: 일별 감소 비율 (현재 미사용, 확장용)
        days: 프로젝션 기간
        launch_period: 런칭 마케팅 집중 기간 (기본 30일)
        sustaining_ratio: 런칭 후 유지 NRU 비율 (기본 10%)
    
    Returns:
        일별 NRU 리스트
    """
    nru_series = []
    
    # 🔥 핵심 수정: Area Normalization
    # Step 1: 런칭 기간 NRU 패턴 생성 (Power Law Decay: 1/t^0.8)
    nru_decay_pattern = []
    for t in range(1, launch_period + 1):
        decay_value = 1.0 / (t ** 0.8)  # D1=1.0, D2=0.57, D3=0.44, ...
        nru_decay_pattern.append(decay_value)
    
    # Step 2: 패턴의 면적(Area) 계산 - 총량 보존의 법칙!
    pattern_area = sum(nru_decay_pattern)
    
    # Step 3: D1 Scale Factor = 총 유저 수 / 패턴 면적
    # 이렇게 하면 런칭 기간 NRU의 합 = total_nru가 됨
    d1_scale = total_nru / pattern_area if pattern_area > 0 else 0
    
    # Phase 1: 런칭 기간 (D1~D30) - 정규화된 패턴 적용
    for day in range(min(launch_period, days)):
        # 정규화된 NRU = Scale × 패턴값
        daily_nru = int(d1_scale * nru_decay_pattern[day])
        nru_series.append(max(daily_nru, 10))  # 최소값 10으로 설정
    
    # Phase 2: 런칭 후 유지 기간 (D31~D365)
    # D30의 NRU를 기준으로 sustaining_ratio만큼 유지
    d30_nru = nru_series[-1] if nru_series else 100
    sustaining_nru = int(d30_nru * sustaining_ratio * 10)  # D30의 ~100% 수준에서 시작
    
    for day in range(launch_period, days):
        # 유지 기간에도 서서히 감소 (월 5% 감소)
        months_after_launch = (day - launch_period) / 30
        decay = np.exp(-0.05 * months_after_launch)
        daily_nru = int(sustaining_nru * decay)
        nru_series.append(max(daily_nru, 10))
    
    return nru_series[:days]


# ============================================
# V8.5: UA/Brand 분리 NRU 계산 (Organic Boost)
# ============================================
def calculate_organic_boost(brand_budget: int, ua_budget: int) -> float:
    """
    브랜딩 예산에 따른 Organic Ratio 증폭 계수 계산
    
    로직:
    - brand_budget이 ua_budget의 0%일 때: 1.0배 (증폭 없음)
    - brand_budget이 ua_budget의 50%일 때: 1.5배
    - brand_budget이 ua_budget의 100%일 때: 2.0배
    - brand_budget이 ua_budget의 200%일 때: 2.5배 (수확체감)
    
    Logarithmic 함수를 사용해 수확체감 효과 적용
    """
    if ua_budget <= 0:
        return 1.0
    
    ratio = brand_budget / ua_budget
    # Logarithmic boost: 1 + ln(1 + ratio) * 0.7
    # ratio=0.5 → 1.28배, ratio=1.0 → 1.49배, ratio=2.0 → 1.77배
    boost = 1.0 + np.log(1 + ratio) * 0.7
    return min(boost, 3.0)  # 최대 3배로 캡


def generate_nru_series_v85(
    ua_budget: int,
    brand_budget: int, 
    target_cpa: int,
    base_organic_ratio: float,
    days: int = 365,
    launch_period: int = 30,
    sustaining_budget_monthly: int = 0,
    # V8.5+ 신규 파라미터
    pre_marketing_ratio: float = 0.0,        # 사전 마케팅 비중
    wishlist_conversion_rate: float = 0.15,  # 위시리스트 전환율
    cpa_saturation_enabled: bool = True,     # CPA 포화 효과
    brand_time_lag_enabled: bool = True      # 브랜딩 지연 효과
) -> tuple:
    """
    V8.5+ NRU 시리즈 생성 - UA/Brand 분리 + Pre-Launch + CPA Saturation
    
    🔥 핵심 로직:
    1. CPA Saturation: 예산 규모에 따라 CPA 상승 (시장 포화 효과)
    2. Pre-Launch Reservoir: 사전예약/위시리스트 유저를 D1에 폭발적 유입
    3. Brand Time-Lag: 브랜딩 효과가 서서히 나타나고 잔존
    
    Args:
        ua_budget: 퍼포먼스 마케팅 예산 (직접 유입)
        brand_budget: 브랜딩 예산 (Organic Boost)
        target_cpa: CPI/CPA 단가
        base_organic_ratio: 기본 자연 유입 비율
        days: 프로젝션 기간
        launch_period: 런칭 마케팅 집중 기간
        sustaining_budget_monthly: 월간 유지 마케팅 예산
        pre_marketing_ratio: 사전 마케팅 비중 (0~1)
        wishlist_conversion_rate: 위시리스트/사전예약 전환율
        cpa_saturation_enabled: CPA 상승 계수 활성화
        brand_time_lag_enabled: 브랜딩 지연 효과 활성화
    
    Returns:
        (nru_series, paid_nru_total, organic_nru_total, organic_boost, meta_info)
    """
    import math
    
    # ============================================
    # 1. CPA Saturation Effect (시장 포화)
    # ============================================
    # 예산이 클수록 효율 좋은 유저가 고갈되어 CPA 상승
    # 공식: Effective CPA = Target CPA × (1 + (Budget / 5억) × 0.05)
    if cpa_saturation_enabled and ua_budget > 0:
        saturation_factor = 1 + (ua_budget / 500_000_000) * 0.05
        effective_cpa = int(target_cpa * saturation_factor)
    else:
        saturation_factor = 1.0
        effective_cpa = target_cpa
    
    # ============================================
    # 2. UA/Brand 예산 분리 및 NRU 계산
    # ============================================
    # 2-1. Pre-Launch 예산과 Post-Launch 예산 분리
    pre_launch_ua = int(ua_budget * pre_marketing_ratio)
    post_launch_ua = ua_budget - pre_launch_ua
    
    # 2-2. Paid NRU 계산 (Effective CPA 적용)
    pre_launch_paid_nru = pre_launch_ua // effective_cpa if effective_cpa > 0 else 0
    post_launch_paid_nru = post_launch_ua // effective_cpa if effective_cpa > 0 else 0
    
    # 2-3. Organic Boost Factor 계산 (Brand Budget 기반)
    organic_boost = calculate_organic_boost(brand_budget, ua_budget)
    
    # 2-4. Organic NRU 계산
    total_paid_nru = pre_launch_paid_nru + post_launch_paid_nru
    effective_organic_ratio = base_organic_ratio * organic_boost
    organic_nru_total = int(total_paid_nru * effective_organic_ratio)
    
    # ============================================
    # 3. Pre-Launch Reservoir (사전예약/위시리스트)
    # ============================================
    # 사전 마케팅으로 모은 유저 = "저수지"에 담아뒀다가 D1에 터뜨림
    # 위시리스트 전환율 적용 (PC: 10~20%, Mobile: 15~25%)
    wishlist_users = int(pre_launch_paid_nru / wishlist_conversion_rate) if wishlist_conversion_rate > 0 else 0
    d1_burst_users = int(wishlist_users * wishlist_conversion_rate)  # 실제 D1 유입
    
    # D1~D3 버스트 분배: D1=80%, D2=10%, D3=10%
    burst_distribution = [0.80, 0.10, 0.10]
    
    # ============================================
    # 4. Brand Time-Lag Effect (브랜딩 지연 효과)
    # ============================================
    # 브랜딩 효과는 Bell Curve로 서서히 나타나고 잔존
    # D-30 ~ D+60 구간에 정규분포로 분산
    brand_effect_curve = []
    if brand_time_lag_enabled and brand_budget > 0:
        # 정규분포 (평균=15, 표준편차=20) → D1~D60 구간에 효과 분포
        for day in range(days):
            # Bell curve centered at D15 with spread of 20 days
            effect = math.exp(-0.5 * ((day - 15) / 20) ** 2)
            brand_effect_curve.append(effect)
        # 정규화
        total_effect = sum(brand_effect_curve)
        brand_effect_curve = [e / total_effect for e in brand_effect_curve] if total_effect > 0 else [0] * days
    else:
        # Time-Lag 비활성화 시 즉시 효과
        brand_effect_curve = [1.0 / 30 if i < 30 else 0 for i in range(days)]
    
    # ============================================
    # 5. NRU 시리즈 생성 (통합)
    # ============================================
    nru_series = [0] * days
    
    # 5-1. Pre-Launch Burst (D1~D3 폭발)
    for i, ratio in enumerate(burst_distribution):
        if i < days:
            nru_series[i] += int(d1_burst_users * ratio)
    
    # 5-2. Post-Launch UA (런칭 후 퍼포먼스 마케팅)
    # Area Normalization으로 30일간 분배
    nru_decay_pattern = [1.0 / (t ** 0.8) for t in range(1, launch_period + 1)]
    pattern_area = sum(nru_decay_pattern)
    d1_scale = post_launch_paid_nru / pattern_area if pattern_area > 0 else 0
    
    for day in range(min(launch_period, days)):
        daily_nru = int(d1_scale * nru_decay_pattern[day])
        nru_series[day] += max(daily_nru, 0)
    
    # 5-3. Organic NRU (Brand Time-Lag 적용)
    for day in range(days):
        organic_daily = int(organic_nru_total * brand_effect_curve[day])
        nru_series[day] += organic_daily
    
    # 5-4. Sustaining 기간 (D31~D365)
    # [FIX] Sustaining은 비용으로만 처리, NRU는 최소한으로 유지
    # 월 매출의 7%를 Sustaining에 쓰지만, 이는 ROAS 계산에만 반영
    # 실제 NRU는 자연 감쇠 (D30 대비 급격히 감소)
    d30_nru = nru_series[29] if len(nru_series) > 29 else 100
    
    # Sustaining NRU는 D30의 5% 수준에서 시작, 빠르게 감쇠
    base_sustaining_nru = int(d30_nru * 0.05)  # D30의 5% (기존 20%에서 크게 축소)
    
    for day in range(launch_period, days):
        months_after_launch = (day - launch_period) / 30
        # [FIX] 더 가파른 감쇠율 적용 (월 10% 감소 → 6개월 후 ~53%, 12개월 후 ~28%)
        decay = np.exp(-0.1 * months_after_launch)
        daily_nru = int(base_sustaining_nru * decay)
        nru_series[day] += max(daily_nru, 5)  # 최소값 10 → 5로 축소
    
    # 최소값 보장 (5명 이하로 떨어지지 않음)
    nru_series = [max(nru, 5) for nru in nru_series]
    
    # ============================================
    # 6. 메타 정보 반환
    # ============================================
    meta_info = {
        "effective_cpa": effective_cpa,
        "cpa_saturation_factor": round(saturation_factor, 3),
        "pre_launch_users": pre_launch_paid_nru,
        "wishlist_users": wishlist_users,
        "d1_burst_users": d1_burst_users,
        "post_launch_paid_nru": post_launch_paid_nru,
        "organic_boost_factor": round(organic_boost, 2),
        "brand_time_lag_peak_day": 15 if brand_time_lag_enabled else 1
    }
    
    return nru_series[:days], total_paid_nru, organic_nru_total, organic_boost, meta_info

def calculate_pr_pattern(selected_games: List[str], raw_data: dict):
    pr_games = raw_data['games']['payment_rate']
    
    valid_games = [g for g in selected_games if g in pr_games]
    if not valid_games:
        return [0.02] * 365
    
    min_len = min(len(pr_games[g]) for g in valid_games)
    min_len = min(min_len, 365)
    
    pattern = []
    for day in range(min_len):
        day_values = [pr_games[g][day] for g in valid_games if day < len(pr_games[g]) and pr_games[g][day] > 0]
        avg_pr = np.mean(day_values) if day_values else 0.02
        pattern.append(max(avg_pr, 0.001))
    
    while len(pattern) < 365:
        pattern.append(pattern[-1] if pattern else 0.02)
    
    return pattern[:365]

def calculate_arppu_pattern(selected_games: List[str], raw_data: dict):
    arppu_games = raw_data['games']['arppu']
    
    valid_games = [g for g in selected_games if g in arppu_games]
    if not valid_games:
        return [50000] * 365
    
    min_len = min(len(arppu_games[g]) for g in valid_games)
    min_len = min(min_len, 365)
    
    pattern = []
    for day in range(min_len):
        day_values = [arppu_games[g][day] for g in valid_games if day < len(arppu_games[g]) and arppu_games[g][day] > 0]
        avg_arppu = np.mean(day_values) if day_values else 50000
        pattern.append(max(avg_arppu, 1000))
    
    while len(pattern) < 365:
        pattern.append(pattern[-1] if pattern else 50000)
    
    return pattern[:365]

def calculate_dau_matrix(nru_series: List[int], retention_curve: List[float], days: int = 365):
    """
    [R1 Fix] DAU 코호트 계산
    - D0 (설치 당일): 리텐션 = 1.0 (100%)
    - D1 이후: retention_curve[days_since_install - 1]
    """
    daily_dau = []
    
    for active_day in range(days):
        total_dau = 0
        for cohort_day in range(active_day + 1):
            days_since_install = active_day - cohort_day
            
            if cohort_day < len(nru_series):
                nru = nru_series[cohort_day]
                
                if days_since_install == 0:
                    retention = 1.0
                else:
                    idx = days_since_install - 1
                    retention = retention_curve[idx] if idx < len(retention_curve) else 0
                
                total_dau += nru * retention
        daily_dau.append(int(total_dau))
    
    return daily_dau

def calculate_revenue(dau: List[float], pr: List[float], arppu: List[float]):
    """
    일별 매출 계산
    
    Revenue = DAU × PR × (ARPPU / 30)
    
    주의: ARPPU는 '월간' 결제자당 평균 결제액이므로,
          일별 계산 시 30으로 나눠야 함
    """
    revenue = []
    for i in range(len(dau)):
        pr_val = pr[i] if i < len(pr) else pr[-1]
        arppu_val = arppu[i] if i < len(arppu) else arppu[-1]
        
        # ARPPU를 일별로 환산 (월간 ARPPU / 30)
        daily_arppu = arppu_val / 30
        
        # 일별 매출 = DAU × PR × 일별 ARPPU
        daily_revenue = dau[i] * pr_val * daily_arppu
        revenue.append(daily_revenue)
    
    return revenue

# Claude AI Integration
# V9.8: 안전한 모델명 설정 (실제 존재하는 모델)
CURRENT_MODEL = "claude-3-5-sonnet-20240620"  # ✅ 실제 작동하는 최신 모델

async def get_claude_insight(prompt: str) -> str:
    """Call Claude API for AI insights with Mock Fallback"""
    if not CLAUDE_API_KEY:
        print("💡 API Key가 없습니다. Mock 데이터를 반환합니다.")
        return None  # Mock으로 폴백
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(
                CLAUDE_API_URL,
                headers={
                    "x-api-key": CLAUDE_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json"
                },
                json={
                    "model": CURRENT_MODEL,  # ✅ 올바른 모델명 사용
                    "max_tokens": 2000,
                    "messages": [
                        {"role": "user", "content": prompt}
                    ]
                }
            )
            
            # API 호출 실패 시 예외 발생
            response.raise_for_status()
            
            data = response.json()
            return data["content"][0]["text"]
                
    except httpx.HTTPStatusError as e:
        print(f"❌ API HTTP 에러: {e.response.status_code}")
        return None
    except Exception as e:
        print(f"❌ AI 호출 에러: {str(e)}")
        print("🔄 안전하게 Mock 데이터로 전환합니다.")
        return None

def create_insight_prompt(summary: Dict[str, Any], analysis_type: str) -> str:
    """Create prompt for Claude based on analysis type with Multi-Persona approach"""
    
    # V7 설정 정보 추출
    v7_settings = summary.get('v7_settings', {})
    blending = summary.get('blending', {})
    
    # V9.2: 플랫폼별 용어 동적 설정
    platforms = blending.get('platforms', ['PC'])
    cost_metric = "CPI" if "Mobile" in platforms else "CPA"
    is_pc_console = any(p in ['PC', 'Console'] for p in platforms)
    
    # V9.2: BEP 상태 계산
    bep_day = summary.get('bep_day', -1)
    bep_status = ""
    if bep_day <= 0:
        bep_status = f"""
[⚠️ Critical Issue: BEP 미달성]
현재 구조로는 1년 내 투자 회수가 어렵습니다. 분석 시 다음 전략을 반드시 포함하세요:
1. {cost_metric} 절감 방안: 타겟팅 최적화 또는 오가닉 비중 확대
2. LTV 개선: 리텐션 D30을 5%p 올리거나 ARPPU를 15% 상향하는 시뮬레이션 제안
3. BM 재검토: 패키지 가격 또는 인게임 결제 모델 조정"""
    else:
        bep_status = f"BEP는 D+{bep_day}에 달성될 것으로 예상됩니다. 안정적인 현금 흐름이 기대됩니다."
    
    base_context = f"""당신은 게임 KPI 프로젝션 분석을 수행하는 4명의 전문가 패널입니다.

[전문가 패널 구성]
1. UA 및 브랜딩 마케터 전문가: {cost_metric} 적정성, 모객 효율, UA 전략, CAC/LTV 분석, Organic Boost 평가
2. 데이터 사이언스 전문가: 지표 건전성, 리텐션 패턴, 통계적 신뢰도, 예측 정확도
3. 퍼블리싱 전문가: BM 구조, 시장 경쟁력, 장르 특성, 글로벌 트렌드, 런칭 타이밍
4. 라이브 서비스 전문가: BEP, ROAS, 투자 회수, Sustaining 전략, 콘텐츠 운영

[플랫폼 컨텍스트]
- 플랫폼: {', '.join(platforms)}
- 비용 지표: {cost_metric} ({'PC/Console은 설치당 비용이 아닌 전환당 비용 기준' if is_pc_console else '모바일 설치당 비용 기준'})
{'- 참고: PC/Console 플랫폼은 CPI 기반 UA가 제한적이므로 Steam 노출, 미디어 리뷰, 커뮤니티 바이럴 등 Organic 중심으로 평가하세요.' if is_pc_console else ''}

[BEP 상태]
{bep_status}

[프로젝션 결과 요약]
프로젝션 기간: {summary.get('projection_days', 365)}일
런칭일: {summary.get('launch_date', 'N/A')}

Best 시나리오:
- 총 Gross Revenue: {summary.get('best', {}).get('gross_revenue', 0):,.0f}원
- 총 NRU: {summary.get('best', {}).get('total_nru', 0):,}명
- Peak DAU: {summary.get('best', {}).get('peak_dau', 0):,}명
- 평균 DAU: {summary.get('best', {}).get('average_dau', 0):,}명

Normal 시나리오:
- 총 Gross Revenue: {summary.get('normal', {}).get('gross_revenue', 0):,.0f}원
- 총 NRU: {summary.get('normal', {}).get('total_nru', 0):,}명
- Peak DAU: {summary.get('normal', {}).get('peak_dau', 0):,}명
- 평균 DAU: {summary.get('normal', {}).get('average_dau', 0):,}명

Worst 시나리오:
- 총 Gross Revenue: {summary.get('worst', {}).get('gross_revenue', 0):,.0f}원
- 총 NRU: {summary.get('worst', {}).get('total_nru', 0):,}명
- Peak DAU: {summary.get('worst', {}).get('peak_dau', 0):,}명
- 평균 DAU: {summary.get('worst', {}).get('average_dau', 0):,}명

[V7 산술 근거 - 이 결과가 어떻게 도출되었는지]
- 블렌딩 비율: 내부 표본 {blending.get('weight_internal', 0.7)*100:.0f}% + 벤치마크 {blending.get('weight_benchmark', 0.3)*100:.0f}%
- Time-Decay: {blending.get('time_decay', True)} (D1:내부90% → D365:벤치마크90%)
- 품질 등급: {v7_settings.get('quality_score', 'B')}급 (승수 ×{v7_settings.get('quality_multiplier', 1.0)})
- BM 타입: {v7_settings.get('bm_type', 'Midcore')}
- 적용 지역: {', '.join(v7_settings.get('regions', ['global']))}
- 계절성 적용: {v7_settings.get('seasonality_applied', True)}
- 벤치마크 기준: {blending.get('genre', 'N/A')} / {', '.join(platforms)}

[중요 지시사항]
- 마크다운 문법(###, **, -, * 등)을 절대 사용하지 마세요
- 일반 텍스트로만 작성하세요
- 번호는 1. 2. 3. 형식으로 사용하세요
- 강조는 따옴표나 괄호로 표현하세요
- 의사결정 지원용으로 전문적이고 간결하게 작성하세요
- 장르 컨텍스트를 정확히 반영하세요 (입력된 장르: {blending.get('genre', 'N/A')})
- BEP 미달성 시 반드시 개선 전략을 포함하세요
"""
    
    type_prompts = {
        "executive_report": f"""
[분석 요청: 종합분석 보고서]
4명의 전문가가 각자의 관점에서 분석하고, 최종 의사결정을 위한 종합 보고서를 작성해주세요.

플랫폼: {', '.join(blending.get('platforms', ['PC']))}
{'- PC/Console 플랫폼: CPI/CPA 기반 UA가 제한적이므로 Steam 노출, 미디어 리뷰, 커뮤니티 바이럴 등 Organic 중심으로 평가하세요.' if any(p in ['PC', 'Console'] for p in blending.get('platforms', ['PC'])) else ''}

응답 형식:

[1. Executive Summary - 핵심 요약]
Normal 시나리오 기준 1년 예상 매출과 핵심 지표를 한 문장으로 요약

[2. 산술 근거 및 가정]
- 이 프로젝션이 어떤 가정과 로직으로 도출되었는지 설명
- 블렌딩 비율, 품질 등급, BM 타입이 결과에 미친 영향

[3. 전문가 통합 분석]
UA&브랜딩 마케터, 퍼블리싱, 데이터 사이언스, 라이브 서비스 4명의 전문가 관점을 종합하여 다음 사항을 하나의 통합된 분석으로 작성:
- 모객 효율 및 {'Organic 중심 마케팅 전략' if any(p in ['PC', 'Console'] for p in blending.get('platforms', ['PC'])) else 'UA 전략'}
- {blending.get('genre', 'N/A')} 장르 시장 경쟁력 및 BM 구조 적합성
- 리텐션 커브 건전성 및 Best-Worst 편차 ({((summary.get('best', {{}}).get('gross_revenue', 1) / max(summary.get('worst', {{}}).get('gross_revenue', 1), 1) - 1) * 100):.0f}%)
- ROAS, 손익분기점, Sustaining 전략

각 전문가의 의견을 나열하지 말고, 하나의 통합된 문단으로 자연스럽게 연결하여 작성하세요.

[4. 프로젝션 신뢰도 평가]
- 데이터 신뢰도: (표본 수, 벤치마크 정합성)
- 가정의 현실성: ({'Organic 비율' if any(p in ['PC', 'Console'] for p in blending.get('platforms', ['PC'])) else 'CPI'}, 리텐션, ARPU 가정 적정성)
- 편차 분석: Best-Worst 시나리오 간 편차 분석

[5. 리스크 분석 및 BEP 달성 전략]
- 핵심 리스크 3가지와 완화 전략
- BEP(손익분기점) 달성이 어려운 경우: 달성을 위한 구체적 개선 방안 제시 (마케팅 효율화, 리텐션 개선, ARPU 향상 등)

[6. 경쟁력 분석]
- {blending.get('genre', 'N/A')} 장르 시장 내 예상 포지셔닝

[7. Go/No-Go 권고]
- 최종 권고: (Go / Conditional Go / No-Go 중 하나)
- 권고 이유: 한 문장
- 권장 액션 3가지

총 1200자 이내로 작성하세요.
""",
        "general": """
[분석 요청: 종합 분석]
4명의 전문가(UA&브랜딩 마케터, 퍼블리싱, 데이터 사이언스, 라이브 서비스)의 관점을 종합한 통합 분석을 작성해주세요.

응답 형식:
1. 통합 분석: 모객 효율, 시장 경쟁력, 지표 건전성, 운영 및 투자 회수 관점을 하나의 문단으로 통합하여 작성 (각 전문가 의견을 나열하지 말고 자연스럽게 연결)
2. 핵심 강점 2가지
3. 핵심 리스크 2가지
4. 권장 액션 3가지

총 400자 이내로 작성하세요.
""",
        "reliability": f"""
[분석 요청: 신뢰도 평가]
4명의 전문가(UA&브랜딩 마케터, 퍼블리싱, 데이터 사이언스, 라이브 서비스)가 이 프로젝션의 신뢰도를 종합적으로 평가해주세요.

플랫폼: {', '.join(blending.get('platforms', ['PC']))}
{'- PC/Console 플랫폼은 모바일과 달리 CPI/CPA 기반 UA가 제한적이므로, Steam/스토어 노출, 미디어 리뷰, 커뮤니티 바이럴 등 Organic 중심 모객을 기준으로 평가하세요.' if any(p in ['PC', 'Console'] for p in blending.get('platforms', ['PC'])) else '- 모바일 플랫폼은 CPI/CPA 기반 UA 효율을 중심으로 평가하세요.'}

응답 형식:
1. 신뢰도 점수: (100점 만점, 숫자만)
2. 신뢰도 등급: (A/B/C/D/F 중 하나)
3. 통합 신뢰도 평가: {'모객 목표 현실성 (Organic 중심), ' if any(p in ['PC', 'Console'] for p in blending.get('platforms', ['PC'])) else 'NRU/CPI 목표 현실성, '}표본 데이터 품질, 시장 벤치마크 적정성, 수익 예측 현실성을 하나의 통합된 문단으로 분석
4. 신뢰도 향상 제안: 구체적인 개선 방안 3가지

총 500자 이내로 작성하세요.
""",
        "retention": """
[분석 요청: 리텐션 분석]
4명의 전문가(UA&브랜딩 마케터, 퍼블리싱, 데이터 사이언스, 라이브 서비스)가 리텐션 및 DAU 패턴을 종합적으로 분석해주세요.

응답 형식:
1. DAU 패턴 건강도: (좋음/보통/우려 중 하나와 이유)
2. 통합 리텐션 분석: 리텐션 커브 분석, 장르 대비 수준, UA 효율 영향을 하나의 통합된 문단으로 분석 (각 전문가 의견을 나열하지 말 것)
3. 리텐션 개선 액션 플랜: 우선순위별 3가지

총 400자 이내로 작성하세요.
""",
        "revenue": """
[분석 요청: 매출 분석]
4명의 전문가(UA&브랜딩 마케터, 퍼블리싱, 데이터 사이언스, 라이브 서비스)가 매출 예측을 종합적으로 분석해주세요.

응답 형식:
1. 매출 예측 현실성: (낙관적/적정/보수적 중 하나와 이유)
2. 통합 매출 분석: 손익분기점, ARPU, 과금 전환율, 시장 점유율을 하나의 통합된 문단으로 분석 (각 전문가 의견을 나열하지 말 것)
3. 매출 극대화 전략: 우선순위별 3가지

총 400자 이내로 작성하세요.
""",
        "risk": """
[분석 요청: 리스크 분석]
4명의 전문가(UA&브랜딩 마케터, 퍼블리싱, 데이터 사이언스, 라이브 서비스)가 리스크 요인을 종합적으로 분석해주세요.

응답 형식:
1. 전체 리스크 수준: (높음/중간/낮음 중 하나)
2. Best-Worst 편차 분석: (편차 비율과 의미)
3. 통합 리스크 분석: 재무 리스크, UA 리스크, 예측 불확실성, 시장/경쟁 리스크를 하나의 통합된 문단으로 분석 (각 전문가 의견을 나열하지 말 것)
4. 리스크 완화 전략: 우선순위별 3가지

총 450자 이내로 작성하세요.
""",
        "competitive": """
[분석 요청: 경쟁력 분석]
4명의 전문가(UA&브랜딩 마케터, 퍼블리싱, 데이터 사이언스, 라이브 서비스)가 시장 경쟁력을 종합적으로 분석해주세요.

응답 형식:
1. 시장 경쟁력 등급: (상/중/하 중 하나와 이유)
2. 통합 경쟁력 분석: 장르 내 포지셔닝, 차별화 포인트, 수익 모델 경쟁력을 하나의 통합된 문단으로 분석 (각 전문가 의견을 나열하지 말 것)
3. 경쟁력 강화 전략: 우선순위별 3가지

총 400자 이내로 작성하세요.
"""
    }
    
    return base_context + type_prompts.get(analysis_type, type_prompts["general"])

# API Endpoints
@app.get("/")
async def root():
    return {"message": "Game KPI Projection API", "version": "2.0.0", "ai_enabled": bool(CLAUDE_API_KEY)}

@app.get("/api/games")
async def get_available_games():
    raw_data = load_raw_data()
    return {
        "retention": list(raw_data['games']['retention'].keys()),
        "nru": list(raw_data['games']['nru'].keys()),
        "payment_rate": list(raw_data['games']['payment_rate'].keys()),
        "arppu": list(raw_data['games']['arppu'].keys())
    }

@app.get("/api/games/metadata")
async def get_games_metadata():
    """Get metadata for all games (release date, genre, platform, etc.)"""
    raw_data = load_raw_data()
    return raw_data.get('game_metadata', {})

@app.get("/api/games/{metric}/{game_name}")
async def get_game_data(metric: str, game_name: str):
    raw_data = load_raw_data()
    
    if metric not in raw_data['games']:
        raise HTTPException(status_code=404, detail=f"Metric '{metric}' not found")
    
    if game_name not in raw_data['games'][metric]:
        raise HTTPException(status_code=404, detail=f"Game '{game_name}' not found in {metric}")
    
    return {
        "game": game_name,
        "metric": metric,
        "data": raw_data['games'][metric][game_name]
    }

@app.get("/api/config")
async def get_default_config():
    return load_config()

@app.post("/api/projection")
async def calculate_projection(input_data: ProjectionInput):
    raw_data = load_raw_data()
    
    days = input_data.projection_days
    results = {"best": {}, "normal": {}, "worst": {}}
    
    # ============================================
    # V7: 블렌딩 설정 추출
    # ============================================
    blending = input_data.blending or {}
    base_weight = blending.get("weight", 0.7)  # 기본값: 내부 70%
    genre = blending.get("genre", "MMORPG")
    platforms = blending.get("platforms", ["PC"])
    use_benchmark_only = blending.get("benchmark_only", False)
    use_time_decay = blending.get("time_decay", True)  # V7: Time-Decay 기본 활성화
    
    # V7: Quality Score & BM Type
    quality_grade = input_data.quality_score or "B"
    quality_multiplier = QUALITY_SCORES.get(quality_grade, 1.0)
    bm_type = input_data.bm_type or "Midcore"
    bm_modifier = BM_TYPE_MODIFIERS.get(bm_type, {"pr_mod": 1.0, "arppu_mod": 1.0})
    
    # V7: 계절성 팩터
    regions = input_data.regions or ["global"]
    seasonality_factors = calculate_seasonality(regions, input_data.launch_date, days)
    
    # 표본 게임이 없으면 벤치마크 100% 사용
    has_sample_games = len(input_data.retention.selected_games) > 0
    if not has_sample_games:
        base_weight = 0.0
        use_benchmark_only = True
    
    # 벤치마크 데이터 가져오기 (BM Type 적용)
    benchmark = get_benchmark_data(genre, platforms)
    benchmark["pr"] = benchmark["pr"] * bm_modifier["pr_mod"]
    benchmark["arppu"] = benchmark["arppu"] * bm_modifier["arppu_mod"]
    benchmark_ret_curve = generate_benchmark_retention_curve(benchmark, days)
    
    # 내부 표본 기반 계수 계산
    a, b = calculate_retention_coefficients(input_data.retention.selected_games, raw_data)
    pr_pattern = calculate_pr_pattern(input_data.revenue.selected_games_pr, raw_data)
    arppu_pattern = calculate_arppu_pattern(input_data.revenue.selected_games_arppu, raw_data)
    
    for scenario in ["best", "normal", "worst"]:
        target_d1 = input_data.retention.target_d1_retention[scenario]
        
        # 내부 표본 기반 리텐션 커브
        internal_ret_curve = generate_retention_curve(a, b, target_d1, days)
        
        # V7: Time-Decay 블렌딩 적용
        if use_time_decay and not use_benchmark_only:
            # 벤치마크 커브를 target_d1에 맞게 스케일링
            benchmark_scale = target_d1 / benchmark["d1"] if benchmark["d1"] > 0 else 1.0
            scaled_benchmark_curve = [min(r * benchmark_scale, 1.0) for r in benchmark_ret_curve]
            ret_curve = calculate_time_decay_blended_retention(
                internal_ret_curve, scaled_benchmark_curve, days, quality_multiplier
            )
        elif not use_benchmark_only:
            # 기존 고정 블렌딩
            benchmark_scale = target_d1 / benchmark["d1"] if benchmark["d1"] > 0 else 1.0
            scaled_benchmark_curve = [min(r * benchmark_scale, 1.0) for r in benchmark_ret_curve]
            ret_curve = calculate_blended_retention(internal_ret_curve, scaled_benchmark_curve, base_weight)
        else:
            # 벤치마크만 사용
            benchmark_scale = target_d1 / benchmark["d1"] if benchmark["d1"] > 0 else 1.0
            ret_curve = [min(r * benchmark_scale * quality_multiplier, 1.0) for r in benchmark_ret_curve]
        
        # V7: NRU 시리즈 생성 (런칭 마케팅 D1~D30 집중)
        d1_nru = input_data.nru.d1_nru[scenario]
        
        # 시나리오별 NRU 보정
        nru_adj = input_data.nru.adjustment.get("best_vs_normal", 0) if scenario == "best" else \
                  input_data.nru.adjustment.get("worst_vs_normal", 0) if scenario == "worst" else 0
        adjusted_d1_nru = int(d1_nru * (1 + nru_adj))
        
        # V8.5: UA/Brand 분리 지원
        ua_budget = input_data.nru.ua_budget or 0
        brand_budget = input_data.nru.brand_budget or 0
        target_cpa = input_data.nru.target_cpa or 2000
        base_organic_ratio = input_data.nru.base_organic_ratio or 0.2
        
        # UA/Brand 예산이 설정되어 있으면 V8.5 로직 사용
        if ua_budget > 0:
            # 시나리오별 예산 조정
            scenario_mult = 1.1 if scenario == "best" else (0.9 if scenario == "worst" else 1.0)
            adj_ua = int(ua_budget * scenario_mult)
            adj_brand = int(brand_budget * scenario_mult)
            
            sustaining_monthly = input_data.basic_settings.get("sustaining_mkt_budget_monthly", 0) if input_data.basic_settings else 0
            
            # V8.5+ 신규 파라미터
            pre_marketing_ratio = input_data.nru.pre_marketing_ratio or 0.0
            wishlist_conversion_rate = input_data.nru.wishlist_conversion_rate or 0.15
            cpa_saturation_enabled = input_data.nru.cpa_saturation_enabled if input_data.nru.cpa_saturation_enabled is not None else True
            brand_time_lag_enabled = input_data.nru.brand_time_lag_enabled if input_data.nru.brand_time_lag_enabled is not None else True
            
            nru_series, paid_nru, organic_nru, organic_boost, nru_meta = generate_nru_series_v85(
                adj_ua, adj_brand, target_cpa, base_organic_ratio, days, 30, sustaining_monthly,
                pre_marketing_ratio, wishlist_conversion_rate, cpa_saturation_enabled, brand_time_lag_enabled
            )
            
            # 시나리오별 메타 정보 저장
            if scenario == "normal":
                v85_nru_meta = {
                    "paid_nru": paid_nru,
                    "organic_nru": organic_nru,
                    "organic_boost_factor": round(organic_boost, 2),
                    "total_nru": paid_nru + organic_nru,
                    # V8.5+ 추가 메타
                    "effective_cpa": nru_meta["effective_cpa"],
                    "cpa_saturation_factor": nru_meta["cpa_saturation_factor"],
                    "pre_launch_users": nru_meta["pre_launch_users"],
                    "wishlist_users": nru_meta["wishlist_users"],
                    "d1_burst_users": nru_meta["d1_burst_users"],
                    "brand_time_lag_peak_day": nru_meta["brand_time_lag_peak_day"]
                }
        else:
            # 기존 로직 (d1_nru 직접 입력)
            nru_series = generate_nru_series(adjusted_d1_nru, [], days)
            v85_nru_meta = None
        
        # V7: 계절성 적용 (NRU에 반영)
        nru_series = [int(nru * sf) for nru, sf in zip(nru_series, seasonality_factors)]
        
        # DAU 계산
        dau_series = calculate_dau_matrix(nru_series, ret_curve, days)
        
        # PR 보정
        pr_adj = input_data.revenue.pr_adjustment.get("best_vs_normal", 0) if scenario == "best" else \
                 input_data.revenue.pr_adjustment.get("worst_vs_normal", 0) if scenario == "worst" else 0
        
        # PR 블렌딩 (BM Type 적용됨) + V7: Quality Score도 적용
        if not use_benchmark_only:
            weight_internal = base_weight
            # 벤치마크 PR에 Quality Score 적용
            adjusted_benchmark_pr = benchmark["pr"] * quality_multiplier
            pr_series = calculate_blended_pr(pr_pattern, adjusted_benchmark_pr, weight_internal, days)
        else:
            # 벤치마크만 사용 시에도 Quality Score 적용
            pr_series = [benchmark["pr"] * quality_multiplier] * days
        pr_series = [p * (1 + pr_adj) for p in pr_series]
        
        # ARPPU 보정
        arppu_adj = input_data.revenue.arppu_adjustment.get("best_vs_normal", 0) if scenario == "best" else \
                    input_data.revenue.arppu_adjustment.get("worst_vs_normal", 0) if scenario == "worst" else 0
        
        # ARPPU 블렌딩 (BM Type 적용됨) + V7: Quality Score도 적용
        if not use_benchmark_only:
            # 벤치마크 ARPPU에 Quality Score 적용
            adjusted_benchmark_arppu = benchmark["arppu"] * quality_multiplier
            arppu_series = calculate_blended_arppu(arppu_pattern, adjusted_benchmark_arppu, base_weight, days)
        else:
            # 벤치마크만 사용 시에도 Quality Score 적용
            arppu_series = [benchmark["arppu"] * quality_multiplier] * days
        arppu_series = [a * (1 + arppu_adj) for a in arppu_series]
        
        # V7: 계절성을 ARPPU에도 반영
        arppu_series = [arppu * sf for arppu, sf in zip(arppu_series, seasonality_factors)]
        
        # Revenue 계산 (일별 ARPPU 환산 적용됨)
        revenue_series = calculate_revenue(dau_series, pr_series, arppu_series)
        
        results[scenario] = {
            "retention": {
                "coefficients": {"a": float(a), "b": float(b)},
                "target_d1": target_d1,
                "curve": ret_curve[:90]
            },
            "nru": {
                "d1_nru": d1_nru,
                "series": nru_series[:90],
                "total": sum(nru_series),
                "paid": paid_nru if ua_budget > 0 else sum(nru_series),
                "organic": organic_nru if ua_budget > 0 else 0
            },
            "dau": {
                "series": dau_series[:90],
                "peak": int(max(dau_series)),
                "average": int(np.mean(dau_series))
            },
            "revenue": {
                "pr_series": pr_series[:90],
                "arppu_series": arppu_series[:90],
                "daily_revenue": revenue_series[:90],
                "total_gross": sum(revenue_series),
                "average_daily": float(np.mean(revenue_series))
            },
            "full_data": {
                "nru": nru_series,
                "dau": dau_series,
                "revenue": revenue_series,
                "retention": ret_curve,
                "pr": pr_series,
                "arppu": arppu_series
            }
        }
    
    # Calculate summary
    summary = {}
    
    # V8.5: 마케팅 예산 총합 계산
    ua_budget = input_data.nru.ua_budget or 0
    brand_budget = input_data.nru.brand_budget or 0
    basic = input_data.basic_settings or load_config()["basic_settings"]
    sustaining_monthly = basic.get("sustaining_mkt_budget_monthly", 0)
    total_sustaining = sustaining_monthly * 12  # 연간 유지 예산
    
    total_marketing_budget = ua_budget + brand_budget + total_sustaining
    
    for scenario in ["best", "normal", "worst"]:
        gross = results[scenario]["revenue"]["total_gross"]
        
        market_fee = basic.get("market_fee_ratio", 0.3)
        vat = basic.get("vat_ratio", 0.1)
        infra = basic.get("infrastructure_cost_ratio", 0.03)
        
        net = gross * (1 - market_fee - vat - infra)
        
        # V8.5: ROAS 계산 분리
        # Paid ROAS: 퍼포먼스 마케팅(UA) 효율 (마케터용)
        paid_roas = (gross / ua_budget * 100) if ua_budget > 0 else 0
        
        # Blended ROAS: 전체 마케팅 효율 (경영진 보고용)
        blended_roas = (gross / total_marketing_budget * 100) if total_marketing_budget > 0 else 0
        
        # LTV, CAC 계산
        total_nru = results[scenario]["nru"]["total"]
        paid_nru_count = results[scenario]["nru"].get("paid", total_nru)
        
        ltv = gross / total_nru if total_nru > 0 else 0
        cac_paid = ua_budget / paid_nru_count if paid_nru_count > 0 else 0
        cac_blended = total_marketing_budget / total_nru if total_nru > 0 else 0
        
        summary[scenario] = {
            "gross_revenue": gross,
            "net_revenue": net,
            "total_nru": total_nru,
            "peak_dau": results[scenario]["dau"]["peak"],
            "average_dau": results[scenario]["dau"]["average"],
            "average_daily_revenue": results[scenario]["revenue"]["average_daily"],
            # V8.5: ROAS 분리
            "paid_roas": round(paid_roas, 1),      # UA 효율 (마케터용)
            "blended_roas": round(blended_roas, 1), # 전체 효율 (경영진용)
            "ltv": round(ltv, 0),
            "cac_paid": round(cac_paid, 0),
            "cac_blended": round(cac_blended, 0)
        }
    
    # V8.5: 마케팅 예산 분석 정보
    v85_marketing_analysis = {
        "ua_budget": ua_budget,
        "brand_budget": brand_budget,
        "sustaining_budget_annual": total_sustaining,
        "total_marketing_budget": total_marketing_budget,
        "organic_boost_factor": round(calculate_organic_boost(brand_budget, ua_budget), 2) if ua_budget > 0 else 1.0,
        "budget_breakdown": {
            "ua_ratio": round(ua_budget / total_marketing_budget * 100, 1) if total_marketing_budget > 0 else 0,
            "brand_ratio": round(brand_budget / total_marketing_budget * 100, 1) if total_marketing_budget > 0 else 0,
            "sustaining_ratio": round(total_sustaining / total_marketing_budget * 100, 1) if total_marketing_budget > 0 else 0
        },
        # V8.5+ 신규 메타 정보
        "pre_launch_settings": {
            "pre_marketing_ratio": input_data.nru.pre_marketing_ratio or 0.0,
            "wishlist_conversion_rate": input_data.nru.wishlist_conversion_rate or 0.15,
            "cpa_saturation_enabled": input_data.nru.cpa_saturation_enabled if input_data.nru.cpa_saturation_enabled is not None else True,
            "brand_time_lag_enabled": input_data.nru.brand_time_lag_enabled if input_data.nru.brand_time_lag_enabled is not None else True
        },
        "nru_analysis": v85_nru_meta if 'v85_nru_meta' in dir() and v85_nru_meta else None
    }
    
    return {
        "status": "success",
        "input": {
            "launch_date": input_data.launch_date,
            "projection_days": days,
            "retention_games": input_data.retention.selected_games,
            "nru_games": input_data.nru.selected_games,
            "pr_games": input_data.revenue.selected_games_pr,
            "arppu_games": input_data.revenue.selected_games_arppu
        },
        "blending": {
            "weight_internal": base_weight,
            "weight_benchmark": 1 - base_weight,
            "time_decay": use_time_decay,
            "genre": genre,
            "platforms": platforms,
            "benchmark_only": use_benchmark_only,
            "benchmark_data": benchmark
        },
        "v7_settings": {
            "quality_score": quality_grade,
            "quality_multiplier": quality_multiplier,
            "bm_type": bm_type,
            "bm_modifier": bm_modifier,
            "regions": regions,
            "seasonality_applied": True
        },
        "v85_marketing": v85_marketing_analysis,  # V8.5: 마케팅 분석 추가
        "summary": summary,
        "results": results
    }

# V9.8: Mock AI Report Generator (Fallback용)
def generate_mock_ai_report(summary: Dict[str, Any], analysis_type: str) -> str:
    """API 실패 시 사용할 Mock 보고서 생성"""
    genre = summary.get('blending', {}).get('genre', 'N/A')
    platforms = ', '.join(summary.get('blending', {}).get('platforms', ['PC']))
    normal_revenue = summary.get('normal', {}).get('gross_revenue', 0)
    bep_day = summary.get('bep_day', -1)
    
    bep_status = f"D+{bep_day}에 BEP 달성 예상" if bep_day > 0 else "1년 내 BEP 미달성 위험"
    
    if analysis_type == "executive_report":
        return f"""[종합 분석 요약]
{genre} 장르의 {platforms} 플랫폼 프로젝트입니다. 
Normal 시나리오 기준 총 매출 {normal_revenue:,.0f}원이 예상됩니다.
{bep_status}입니다.

[핵심 지표 평가]
1. 매출 전망: Normal 시나리오 기준 적정 수준
2. 리텐션: 장르 평균 대비 검토 필요
3. 마케팅 효율: CPA/CPI 최적화 여지 존재

[리스크 분석]
1. 시장 경쟁: 동일 장르 출시작 모니터링 필요
2. 유저 확보: 런칭 초기 집중 마케팅 권장
3. 수익화: BM 모델 최적화 검토

[전략 제언]
1. 런칭 전 사전 마케팅으로 위시리스트 확보
2. D1 리텐션 확보를 위한 온보딩 최적화
3. 라이브 서비스 준비로 장기 운영 대비

* 이 보고서는 AI 연결 실패로 인한 기본 분석입니다."""
    else:
        return f"[{analysis_type}] {genre} 프로젝트 분석 결과입니다. 상세 AI 분석을 위해 API 연결을 확인해주세요."

# AI Insight Endpoint
@app.post("/api/ai/insight")
async def get_ai_insight_endpoint(request: AIInsightRequest):
    """Get AI-powered insights for projection results with Mock Fallback"""
    prompt = create_insight_prompt(request.projection_summary, request.analysis_type)
    insight = await get_claude_insight(prompt)
    
    # V9.8: Mock Fallback
    if insight is None:
        print("⚠️ AI API failed. Using Mock Report.")
        insight = generate_mock_ai_report(request.projection_summary, request.analysis_type)
        ai_model = "mock-fallback"
    else:
        ai_model = CURRENT_MODEL
    
    return {
        "status": "success",
        "analysis_type": request.analysis_type,
        "insight": insight,
        "ai_model": ai_model
    }

@app.get("/api/ai/status")
async def get_ai_status():
    """Check AI integration status"""
    return {
        "enabled": bool(CLAUDE_API_KEY),
        "model": CURRENT_MODEL,
        "available_types": ["general", "reliability", "retention", "revenue", "risk", "competitive"]
    }

@app.get("/api/raw-data")
async def get_raw_data():
    return load_raw_data()

@app.get("/api/raw-data/download")
async def download_raw_data_excel():
    """Download raw game data as Excel file (same format as original)"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from fastapi.responses import StreamingResponse
    
    raw_data = load_raw_data()
    wb = Workbook()
    
    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def create_raw_sheet(ws, sheet_title, metric_name, description, data_dict):
        """Raw 데이터 시트 생성 (원본 엑셀 형식)"""
        # Row 1: 안내 문구
        ws['B1'] = f'- 아래 게임 추가 시 {sheet_title} 게임 리스트에 자동으로 추가됩니다.'
        ws['B1'].font = Font(color="FF0000")
        
        # Row 2: 메트릭명 및 설명
        ws['B2'] = metric_name
        ws['B2'].font = Font(bold=True)
        ws['C2'] = description
        
        # Row 3: 헤더 (게임명, 1, 2, 3, ... 365)
        ws['B3'] = '게임명'
        ws['B3'].fill = header_fill
        ws['B3'].font = header_font
        ws['B3'].border = thin_border
        
        max_days = 90 if metric_name == '리텐션' else 365
        for day in range(1, max_days + 1):
            col = day + 2  # C부터 시작
            cell = ws.cell(row=3, column=col, value=day)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        # Row 4+: 게임 데이터
        row_idx = 4
        for game_name, values in data_dict.items():
            ws.cell(row=row_idx, column=2, value=game_name).border = thin_border
            for i, val in enumerate(values[:max_days]):
                cell = ws.cell(row=row_idx, column=i + 3, value=val)
                cell.border = thin_border
                if metric_name in ['리텐션', 'PR']:
                    cell.number_format = '0.00%'
            row_idx += 1
        
        # 열 너비 조정
        ws.column_dimensions['B'].width = 20
        for col in range(3, max_days + 3):
            ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = 8
    
    # Raw_Retention 시트
    ws_retention = wb.active
    ws_retention.title = "Raw_Retention"
    create_raw_sheet(ws_retention, "1. Retention", "리텐션", "론칭 ~ 90일까지의 리텐션 정보 입력", raw_data['games'].get('retention', {}))
    
    # Raw_NRU 시트
    ws_nru = wb.create_sheet("Raw_NRU")
    create_raw_sheet(ws_nru, "2. NRU", "NRU", "론칭 ~ 365일까지의 데이터 입력", raw_data['games'].get('nru', {}))
    
    # Raw_PR 시트
    ws_pr = wb.create_sheet("Raw_PR")
    create_raw_sheet(ws_pr, "3. Revenue", "PR", "론칭 ~ 365일까지의 데이터 입력", raw_data['games'].get('payment_rate', {}))
    
    # Raw_ARPPU 시트
    ws_arppu = wb.create_sheet("Raw_ARPPU")
    create_raw_sheet(ws_arppu, "3. Revenue", "ARPPU", "론칭 ~ 365일까지의 데이터 입력", raw_data['games'].get('arppu', {}))
    
    # 메모리에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=raw_game_data.xlsx"}
    )

@app.post("/api/raw-data/upload")
async def upload_game_data(file: UploadFile = File(...), metric: str = "retention"):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are supported")
    
    import pandas as pd
    from io import StringIO
    
    content = await file.read()
    df = pd.read_csv(StringIO(content.decode('utf-8')))
    
    raw_data = load_raw_data()
    
    for _, row in df.iterrows():
        game_name = row.iloc[0]
        values = row.iloc[1:].tolist()
        values = [float(v) for v in values if pd.notna(v)]
        
        if metric in raw_data['games']:
            raw_data['games'][metric][game_name] = values
    
    raw_data['metadata'][f'{metric}_games'] = list(raw_data['games'][metric].keys())
    
    with open(RAW_DATA_PATH, 'w', encoding='utf-8') as f:
        json.dump(raw_data, f, ensure_ascii=False, indent=2)
    
    return {"status": "success", "message": f"Added/updated games in {metric}"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
